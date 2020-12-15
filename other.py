from openpyxl.styles import Color, PatternFill, Font, Border
from bs4 import BeautifulSoup
import openpyxl as xl
import requests
# import threaded
import schedule
import datetime
import time
import re
import os
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.webdriver.chrome.options import Options
options = webdriver.ChromeOptions()
# options.add_argument('--headless')
# options.add_argument("--disable-notifications")

os.system('clear')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
Dir_Name = os.path.join(BASE_DIR, 'TVShowsPythonScraper')
xl_file_path = os.path.join(Dir_Name, 'Shows.xlsx')
min_rating = 7
allowed_channels = ['hbo', 'netflix', 'hulu_plus', 'starz', 'showtime', 'apple_plus'] 


webdriverPath = os.path.join(Dir_Name, 'chromedriver.exe')

driver = webdriver.Chrome(executable_path=webdriverPath, options=options)
test_condition = 2

def ReadExcel():
    wb_obj = xl.load_workbook(xl_file_path)
    sheet_obj = wb_obj.active
    number_of_rows = sheet_obj.max_row
    last_row_index_with_data = GetLastRowIndexWithData(sheet_obj, number_of_rows) 
    # last_row_index_with_date = last row in the excel file with date 
    ratings_from_xl_file = []
    titles_from_xl_file = []
    mod_date_from_xl_file = []

    for i in range(2, last_row_index_with_data+1):
        ratings_from_xl_file.append(sheet_obj.cell(row=i, column=3).value)
        titles_from_xl_file.append(sheet_obj.cell(row=i, column=2).value)
        mod_date_from_xl_file.append(sheet_obj.cell(row=i, column=1).value)

    dictionary = { 
        'titles_from_xl_file':titles_from_xl_file, 
        'ratings_from_xl_file': ratings_from_xl_file, 
        'mod_date_from_xl_file': mod_date_from_xl_file
         }
    return dictionary

def GetLastRowIndexWithData(sheet_obj, number_of_rows ):
    last_row_index_with_data = 0
    while True:
        try:
            # print(sheet_obj.cell(number_of_rows, 3).value)
            if sheet_obj.cell(row=number_of_rows, column=3).value != None:
                last_row_index_with_data = number_of_rows
                break
            elif number_of_rows == 1:
                last_row_index_with_data = number_of_rows
                break
            else:
                number_of_rows -= 1
        except:
            print('something went wrong while reading from excel file')
    return last_row_index_with_data

def GetUnWantedTitles():
    is_last_page = False
    unwanted_titles = []
    
    # my_user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.111 Safari/537.36'
    counter = 0
    # while not is_last_page:
    while counter < test_condition:
        offset = counter*50
        counter += 1
        # url = f'https://reelgood.com/tv?offset={offset}'
        url = f'https://reelgood.com/tv/origin/america?filter-genre=6&filter-genre=39&filter-genre=15&filter-genre=16&filter-genre=18&filter-genre=37&offset={offset}'
        driver.get(url)
        # table_layout = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By., "myDynamicElement"))
        try:
            popup_X = driver.find_element_by_xpath('//div[@data-type="x"]')
            popup_X.click()
        except:
            pass
        
        try:
            table_layout = driver.find_element_by_xpath('//button[@title="Switch to table layout"]')
            table_layout.click()
        except:
            pass
        time.sleep(2)
        try:
            html = driver.page_source
            soup = BeautifulSoup(html, features='html.parser')
            table = soup.find('table', attrs={'class':'css-1179hly'})
        except:
            pass
        data = []
        if table: # checking if the table exists in the current page else the page will be last page 
            print(f'unwanted-titles being scraped at apage # {counter}')
            # tbody = table.find('tbody')
            rows = soup.find_all('tr', attrs={'class':'css-gfsdx9'})
            for row in rows:
                td = row.find_all('td')
                td = [e.text.strip() for e in td]
                data.append([e for e in td if e])
            #  getting unwanted titles
            for row in data:
                unwanted_titles.append(row[0])# appending unwanted titles from website source
            # tds = soup.find_all('td', class_ = lambda value: value == 'css-1vuzpp2')
        else:
            is_last_page = True
    return unwanted_titles

def Process():
    is_last_page = False
    titles = []
    ratings = []
    __available_on = []
    available_on = []
    unwanted_titles = GetUnWantedTitles()
    # print(unwanted_titles)
    # my_user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.111 Safari/537.36'
    counter = 0
    # while not is_last_page:
    while counter < test_condition:
        offset = counter*50
        counter += 1
        # url = f'https://reelgood.com/tv?offset={offset}'
        url = f'https://reelgood.com/tv/origin/america?offset={offset}'
        driver.get(url)
        try:
            popup_X = driver.find_element_by_xpath('//div[@data-type="x"]')
            popup_X.click()# closing the popup
        except:
            pass
        try:
            table_layout = driver.find_element_by_xpath('//button[@title="Switch to table layout"]')
            table_layout.click()# changing the layout to tableview
        except:
            pass
        time.sleep(2)
        try:
            html = driver.page_source
            soup = BeautifulSoup(html, features='html.parser')

            table = soup.find('table', attrs={'class':'css-1179hly'})
        except:
            pass
        data = []
        # print(table)
        if table: # checking if the table exists in the current page else the page will be last page 
            print(f'general titles, ratings and channels are being scrapeed currently at page # {counter}')
            # Netflix, Hulu, HBO, Showtime, Startz, Apple+
           
            rows = soup.find_all('tr', attrs={'class':'css-gfsdx9'})
            for row in rows:
                td = row.find_all('td')
                td = [e.text.strip() for e in td]
                data.append([e for e in td if e])
            #  getting titles and ratings
            for i, row in enumerate(data):
                if row[4].split('/')[0] == 'N': # if a rating is unavailable means N/A
                    rating = 0
                    ratings.append(rating)
                    titles.append(row[0]) 
                else:
                    rating = float(row[4].split('/')[0])
                    titles.append(row[0])# appending titles from website source
                    ratings.append(rating)# appending ratings from website source
                    # getting logos
            tds = soup.find_all('td', class_ = lambda value: value == 'css-1vuzpp2')
            for i, td in enumerate(tds):
                images_in_current_row = []
                images_in_current_row = td.find_all('img')
                logos = []
                for logo in images_in_current_row:
                    logos.append(logo['alt'])
                __available_on.append(logos)
                arr = []
                arr = __available_on[i] # populating arr with current row logo images. 
                temp = ''
                for a in arr:
                    temp += a
                    temp += ','
                available_on.append(temp[:-1]) # removing the last comma
        else:
            is_last_page = True
    
    dictionary = {'titles':titles, 'ratings': ratings, 'available_on': available_on, 'unwanted_titles': unwanted_titles}
    WriteToExcel(dictionary)

def WriteToExcel(dictionary):
    # print(len(dictionary['titles']))
    # print(len(dictionary['ratings']))
    # print(len(dictionary['available_on']))
    dictionary = ApplyFilter(dictionary)
    titles = dictionary['allowed_titles']
    ratings = dictionary['allowed_ratings']
    available_on = dictionary['allowed_available_on']
    
    wb = xl.load_workbook(xl_file_path)
    sheet = wb['Sheet1']
    sheet = wb.active
    ResetExcelSheetColors(sheet)
    dictionary_from_file = ReadExcel()
    titles_from_xl_file = dictionary_from_file['titles_from_xl_file']
    ratings_from_xl_file = dictionary_from_file['ratings_from_xl_file']
    # mod_date_from_xl_file = dictionary_from_file['mod_date_from_xl_file']
    #  creting headings ----------------------------------
    sheet.cell(row=1, column=1).value = "Date Modified"
    sheet.cell(row=1, column=2).value = "Titles"
    sheet.cell(row=1, column=3).value = "Ratings "
    sheet.cell(row=1, column=4).value = "Available On"
    # ----------------------------------------------------
    print(f'titles from website length = {len(titles)}')
    print(f'titles from excel length = {len(titles_from_xl_file)}')
    for i, title in enumerate(titles_from_xl_file):
        if titles_from_xl_file[i] != titles[i]: # checking if any titles has changed at website end.
            sheet.cell(row=i+2, column=1).value = datetime.datetime.now().date()
            # sheet.cell(row=i+2, column=1).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
            sheet.cell(row=i+2, column=2).value = titles[i]
            sheet.cell(row=i+2, column=2).fill = PatternFill(start_color='99FF99', end_color='99FF99',fill_type='solid')
            sheet.cell(row=i+2, column=3).value = ratings[i] # ratings gets the ratings of TV Show
            # sheet.cell(row=i+2, column=3).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
            sheet.cell(row=i+2, column=4).value = available_on[i]
            # sheet.cell(row=i+2, column=4).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
            wb.save(xl_file_path)
        if ratings_from_xl_file[i] != ratings[i]: # checking if any rating has changed at website end.
            sheet.cell(row=i+2, column=1).value = datetime.datetime.now().date()
            sheet.cell(row=i+2, column=1).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
            sheet.cell(row=i+2, column=3).value = ratings[i]
            sheet.cell(row=i+2, column=3).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
            wb.save(xl_file_path)
    wb.save(xl_file_path)
    for i, title in enumerate(titles):
        if ratings[i] >= min_rating:
            sheet.cell(row=i+2, column=1).value = datetime.datetime.now().date()
            sheet.cell(row=i+2, column=2).value = title             # row[0] gets the title of TV Show
            sheet.cell(row=i+2, column=3).value = ratings[i]        # ratings gets the ratings of TV Show
            sheet.cell(row=i+2, column=4).value = available_on[i]
    wb.save(xl_file_path)
    # again reading excel file
    dictionary_from_file = ReadExcel()
    titles_from_xl_file = dictionary_from_file['titles_from_xl_file']
    CreateDirsFromListOfTitlesInExcelFile(titles_from_xl_file)
def CreateDirsFromListOfTitlesInExcelFile(titles_from_xl_file):
    os.chdir('C:\\Users\\Adnan\\OneDrive\\Desktop\\TVShowsPythonScraper')
    # os.chdir('D:/')
    root ='.'
    characters = [':', '*', '?', '\\', '/', '|', '"', '>', '<']
    for title in titles_from_xl_file:
        for c in characters:
            if c in title:
                title = title.replace(c , '_')
        path = f'{root}/TV Shows/{title}'
        # print(path)
        if not os.path.exists(path):
            os.makedirs(path)
    print(f'{len(titles_from_xl_file)} folders created')

def IsAllowed(allowed_channels, ith_available_on):
    for val in allowed_channels:
        if val in ith_available_on:
            return True
    return False

def ResetExcelSheetColors(sheet):
    no_fill = xl.styles.PatternFill(fill_type=None)
    for row in sheet:
        for cell in row:
            cell.fill = no_fill    

def ApplyFilter(dictionary):
    titles = dictionary['titles']
    ratings = dictionary['ratings']
    available_on = dictionary['available_on']
    unwanted_titles = dictionary['unwanted_titles']
    allowed_ratings = []
    allowed_titles = []
    allowed_available_on = []
    for i, title in enumerate(titles):
        if ratings[i] >= min_rating:
            if IsAllowed(allowed_channels, available_on[i]):
                allowed_titles.append(title)
                allowed_ratings.append(ratings[i])
                allowed_available_on.append(available_on[i])
    for i, title in enumerate(allowed_titles):
        for unwanted_title in unwanted_titles:
            if unwanted_title == title:
                print(f'{allowed_titles[i]} is removed ')
                allowed_titles.pop(i)
                allowed_ratings.pop(i)
                allowed_available_on.pop(i)
    return {'allowed_ratings':allowed_ratings, 'allowed_titles': allowed_titles, 'allowed_available_on':allowed_available_on}

def MoveFolders():
    partition_label = "C:\\Users\\Adnan\\OneDrive\\Desktop\\TVShowsPythonScraper"
    current_WD = os.getcwd()
    # partition_label = "D:"
    # os.chdir(os.path.join(partition_label, "\\TV Shows"))
    os.chdir("TV Shows")
    dir_list = [name for name in os.listdir(".") if os.path.isdir(name)]
    # os.chdir(os.path.join(partition_label, "\\New Downloads"))
    os.chdir(current_WD)
    os.chdir("New Downloads")
    new_downloads_list = [name for name in os.listdir(".") if os.path.isdir(name)]
    source = os.path.join(partition_label, "\\New Downloads")
    destination = os.path.join(partition_label, "\\TV Shows")
    
    for name in dir_list:
        # print(name)
        name_parts = name.split(' ')
        # print(name_parts)
        for torrent in new_downloads_list:
            flag_list = []
            for name_part in name_parts:
                # print(name_part, torrent)
                if name_part.lower() in torrent.lower():
                    flag_list.append(1)
            if len(flag_list) == len(name_parts):
                # Move the torrent to the name folder 
                full_destination = os.path.join(destination, name)
                if os.path.exists(full_destination):
                    print('The Folder {full_destination} already exists in the [ {destination} ] location')
                else:
                    print( f' "{torrent}" folder has been moved to "{name}" directory inside ')
                    shutil.move(os.path.join(source, torrent), full_destination)
                    


# starting point .............................. 
Process()
driver.close()
MoveFolders()
print('done')
# schedule.every(1).day.at("01:00").do(Process)
# while 1:
#     schedule.run_pending()
#     time.sleep(1)
