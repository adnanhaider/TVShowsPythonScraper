
from openpyxl.styles import Color, PatternFill, Font, Border
from bs4 import BeautifulSoup
import openpyxl as xl
import requests
# import threaded
import schedule
import datetime
import time
import re
import os

os.system('clear')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
Dir_Name = os.path.join(BASE_DIR, 'TVShowsPythonScraper')
xl_file_path = os.path.join(Dir_Name, 'Shows.xlsx')
min_rating = 7
allowed_channels = ['hbo', 'netflix', 'hulu_plus', 'starz', 'showtime', 'apple_plus'] 

def ReadExcel():
    wb_obj = xl.load_workbook(xl_file_path)
    sheet_obj = wb_obj.active
    number_of_rows = sheet_obj.max_row
    last_row_index_with_data = GetLastRowIndexWithData(sheet_obj, number_of_rows) 
    # last_row_index_with_date = last row in the excel file with date 
    ratings_from_xl_file = []
    titles_from_xl_file = []
    mod_date_from_xl_file = []

    for i in range(2, last_row_index_with_data+1):
        ratings_from_xl_file.append(sheet_obj.cell(row=i, column=3).value)
        titles_from_xl_file.append(sheet_obj.cell(row=i, column=2).value)
        mod_date_from_xl_file.append(sheet_obj.cell(row=i, column=1).value)

    dictionary = { 
        'titles_from_xl_file':titles_from_xl_file, 
        'ratings_from_xl_file': ratings_from_xl_file, 
        'mod_date_from_xl_file': mod_date_from_xl_file
         }
    return dictionary

def GetLastRowIndexWithData(sheet_obj, number_of_rows ):
    last_row_index_with_data = 0
    while True:
        try:
            # print(sheet_obj.cell(number_of_rows, 3).value)
            if sheet_obj.cell(row=number_of_rows, column=3).value != None:
                last_row_index_with_data = number_of_rows
                break
            elif number_of_rows == 1:
                last_row_index_with_data = number_of_rows
                break
            else:
                number_of_rows -= 1
        except:
            print('something went wrong while reading from excel file')
    return last_row_index_with_data

def Process():
    is_last_page = False
    titles = []
    ratings = []
    __available_on = []
    available_on = []
    my_user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.111 Safari/537.36'
    counter = 0
    # while not is_last_page:
    while counter < 5:
        offset = counter*50
        counter += 1
        url = f'https://reelgood.com/tv?offset={offset}'
        header = {"User-Agent": my_user_agent}
        page = requests.get(url, headers=header)
        soup = BeautifulSoup(page.content, 'html.parser')
        time.sleep(5)
        table = soup.find('table', attrs={'class':'css-1179hly'})
        data = []
        if table: # checking if the table exists in the current page else the page will be last page 
            # print(counter)
            # Netflix, Hulu, HBO, Showtime, Startz, Apple+
            tbody = table.find('tbody')
            rows = tbody.find_all('tr')
            for row in rows:
                td = row.find_all('td')
                td = [e.text.strip() for e in td]
                data.append([e for e in td if e])
            #  getting titles and ratings
            for i, row in enumerate(data):
                if row[4].split('/')[0] == 'N': # N/A
                    rating = 0
                    # print(f'inside if of N/A .. raing = {rating}')
                    ratings.append(rating)
                    titles.append(row[0])
                else:
                    rating = float(row[4].split('/')[0])
                    titles.append(row[0])
                    ratings.append(rating)
                    # print(f' rating = {rating}')
                    #  getting logos 
            tds = soup.find_all('td', class_ = lambda value: value == 'css-1vuzpp2')
            for i, td in enumerate(tds):
                images_in_current_row = []
                images_in_current_row = td.find_all('img')
                logos = []
                for logo in images_in_current_row:
                    logos.append(logo['alt'])
                __available_on.append(logos)
                arr = []
                arr = __available_on[i] # populating arr with current row logo images. 
                temp = ''
                for a in arr:
                    temp += a
                    temp += ','
                available_on.append(temp[:-1]) # removing the last comma
        else:
            is_last_page = True
    
    dictionary = {'titles':titles, 'ratings': ratings, 'available_on': available_on }
    WriteToExcel(dictionary)

def WriteToExcel(dictionary):
    dictionary = ApplyFilter(dictionary)
    titles = dictionary['allowed_titles']
    ratings = dictionary['allowed_ratings']
    available_on = dictionary['allowed_available_on']
    wb = xl.load_workbook(xl_file_path)
    sheet = wb['Sheet1']
    sheet = wb.active
    ResetExcelSheetColors(sheet)
    dictionary_from_file = ReadExcel()
    titles_from_xl_file = dictionary_from_file['titles_from_xl_file']
    ratings_from_xl_file = dictionary_from_file['ratings_from_xl_file']
    mod_date_from_xl_file = dictionary_from_file['mod_date_from_xl_file']
    if len(titles_from_xl_file)>2:
        for i, title in enumerate(titles_from_xl_file):
            if ratings_from_xl_file[i] != ratings[i]: # checking if any rating has changed at website end.
                sheet.cell(row=i+2, column=1).value = datetime.datetime.now().date()
                sheet.cell(row=i+2, column=1).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
                sheet.cell(row=i+2, column=3).value = ratings[i]
                sheet.cell(row=i+2, column=3).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
            if titles_from_xl_file[i] != title: # checking if any rating has changed at website end.
                sheet.cell(row=i+2, column=1).value = datetime.datetime.now().date()
                sheet.cell(row=i+2, column=1).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
                sheet.cell(row=i+2, column=2).value = title
                sheet.cell(row=i+2, column=2).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
        wb.save(xl_file_path)
    else:
        sheet.cell(row=1, column=1).value = "Date Modified"
        sheet.cell(row=1, column=2).value = "Titles"
        sheet.cell(row=1, column=3).value = "Ratings "
        sheet.cell(row=1, column=4).value = "Available On"

        for i, title in enumerate(titles):
            if ratings[i] >= min_rating:
                sheet.cell(row=i+2, column=1).value = datetime.datetime.now().date()
                sheet.cell(row=i+2, column=2).value = title             # row[0] gets the title of TV Show
                sheet.cell(row=i+2, column=3).value = ratings[i]        # ratings gets the ratings of TV Show
                sheet.cell(row=i+2, column=4).value = available_on[i]
        wb.save(xl_file_path)
    # CreateDirsFromListOfTitlesInExcelFile(titles_from_xl_file)
def CreateDirsFromListOfTitlesInExcelFile(titles_from_xl_file):
    # os.chdir('C:/')
    root ='.'
    characters = [':', '*', '?', '\\', '/', '|', '"', '>', '<']
    for title in titles_from_xl_file:
        for c in characters:
            if c in title:
                title = title.replace(c , '_')
                print(title)
        path = f'{root}/TV Shows/{title}'
        # print(path)
        if not os.path.exists(path):
            os.makedirs(path)
    print('created')

def IsAllowed(allowed_channels, ith_available_on):
    for val in allowed_channels:
        if val in ith_available_on:
            return True
    return False

def ResetExcelSheetColors(sheet):
    no_fill = xl.styles.PatternFill(fill_type=None)
    for row in sheet:
        for cell in row:
            cell.fill = no_fill    

def ApplyFilter(dictionary):
    titles = dictionary['titles']
    ratings = dictionary['ratings']
    available_on = dictionary['available_on']
    allowed_ratings = []
    allowed_titles = []
    allowed_available_on = []
    for i, title in enumerate(titles):
        if ratings[i] >= min_rating:
            if IsAllowed(allowed_channels, available_on[i]):
                allowed_titles.append(title)
                allowed_ratings.append(ratings[i])
                allowed_available_on.append(available_on[i])
    return {'allowed_ratings':allowed_ratings, 'allowed_titles': allowed_titles, 'allowed_available_on':allowed_available_on}

# starting point .............................. 
Process()
print('done')
# schedule.every(1).day.at("01:00").do(Process)
# while 1:
#     schedule.run_pending()
#     time.sleep(1)
